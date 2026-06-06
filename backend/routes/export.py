"""
数据导出 API — 生成 .xlsx 文件保存到 downloads 目录
"""
import os

from flask import Blueprint, request, jsonify, g

from database.db import get_db, beijing_now
from services.log_service import write_log
from middleware.auth_middleware import login_required

export_bp = Blueprint('export', __name__)


def _data_root():
    """数据根目录：Docker 容器内为 /data，本地通过 HOST_DATA_DIR 或项目根目录推算"""
    import flask
    db_path = flask.current_app.config.get('DATABASE_PATH', '')
    if db_path.startswith('/data/'):
        return '/data'
    host_dir = flask.current_app.config.get('HOST_DATA_DIR', '')
    if host_dir:
        return host_dir
    # 本地开发：项目根目录/data
    return os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'data')


def _download_dir():
    return os.path.join(_data_root(), 'download')


def _host_path(filepath):
    """容器内 Linux 路径 → 宿主机 Windows 路径，用于前端展示"""
    import flask
    host_data_dir = flask.current_app.config.get('HOST_DATA_DIR', '')
    db_path = flask.current_app.config.get('DATABASE_PATH', '')
    # Docker 环境：/data/download/xxx.xlsx → HOST_DATA_DIR\download\xxx.xlsx
    if db_path.startswith('/data/') and host_data_dir:
        rel = filepath[len('/data/'):] if filepath.startswith('/data/') else filepath
        return host_data_dir.rstrip('\\') + '\\' + rel.replace('/', '\\')
    return filepath


@export_bp.route('', methods=['POST'])
@login_required
def export_data():
    data = request.get_json() or {}
    export_type = data.get('type', 'all')
    format_type = data.get('format', 'json')

    db = get_db()

    # 中文列名映射
    CONTRACT_HEADERS = {
        'name': '合同名称', 'category': '类别', 'company': '合同公司',
        'contact_person': '联系人', 'contact_phone': '联系电话', 'contact_email': '联系邮箱',
        'agent': '对接业务员', 'start_date': '开始日期', 'end_date': '到期日期',
        'status': '状态', 'file_path': '本地文件', 'remark': '备注',
        'source': '来源', 'email_reminder': '邮件提醒', 'priority': '重要等级',
        'created_at': '创建时间', 'updated_at': '更新时间',
    }
    PATENT_HEADERS = {
        'name': '专利名称', 'patent_no': '专利号', 'type': '类型',
        'holder': '权利人', 'agent': '对接业务员',
        'application_date': '申请日期', 'expire_date': '到期日期',
        'status': '状态', 'file_path': '本地文件', 'remark': '备注',
        'source': '来源', 'email_reminder': '邮件提醒', 'priority': '重要等级',
        'created_at': '创建时间', 'updated_at': '更新时间',
    }
    INSURANCE_HEADERS = {
        'plate_no': '车牌号', 'brand': '品牌型号', 'insurance_company': '保险公司',
        'insurance_type': '险种', 'amount': '保费金额', 'agent': '对接业务员',
        'start_date': '开始日期', 'end_date': '到期日期',
        'status': '状态', 'file_path': '本地文件', 'remark': '备注',
        'source': '来源', 'email_reminder': '邮件提醒', 'priority': '重要等级',
        'created_at': '创建时间', 'updated_at': '更新时间',
    }

    def _value_map(row, headers):
        """将数据库行转为 {中文列名: 值}"""
        result = {}
        for col, cn_name in headers.items():
            val = row.get(col, '')
            if val is None:
                val = ''
            # 翻译特殊字段
            if col == 'email_reminder':
                val = '是' if val else '否'
            elif col == 'source':
                val = '扫描导入' if val == 'scan' else '手动录入'
            result[cn_name] = val
        return result

    # 构建查询
    def _query(table, headers, create_filter='', exp_filter=''):
        cols = ', '.join(headers.keys())
        if export_type == 'byCreate' and create_filter:
            rows = db.execute(f"SELECT {cols} FROM {table} {create_filter}", _params('create')).fetchall()
        elif export_type == 'byExpire' and exp_filter:
            rows = db.execute(f"SELECT {cols} FROM {table} {exp_filter}", _params('expire')).fetchall()
        else:
            rows = db.execute(f"SELECT {cols} FROM {table}").fetchall()
        return [_value_map(dict(r), headers) for r in rows]

    def _params(kind):
        if kind == 'create':
            return [data.get('startDate', '') + ' 00:00:00', data.get('endDate', '') + ' 23:59:59']
        return [data.get('startDate', ''), data.get('endDate', '')]

    # 校验
    if export_type in ('byCreate', 'byExpire'):
        if not data.get('startDate') or not data.get('endDate'):
            return jsonify({'success': False, 'message': '请选择时间范围'}), 400

    create_where = "WHERE created_at >= ? AND created_at <= ?" if export_type == 'byCreate' else ''
    expire_contract = "WHERE end_date >= ? AND end_date <= ?" if export_type == 'byExpire' else ''
    expire_patent = "WHERE expire_date >= ? AND expire_date <= ?" if export_type == 'byExpire' else ''
    expire_insurance = "WHERE end_date >= ? AND end_date <= ?" if export_type == 'byExpire' else ''

    contracts = _query('contracts', CONTRACT_HEADERS, create_where, expire_contract)
    patents = _query('patents', PATENT_HEADERS, create_where, expire_patent)
    insurances = _query('insurances', INSURANCE_HEADERS, create_where, expire_insurance)

    # JSON 导出（保留兼容）
    if format_type == 'json':
        result = {
            'contracts': contracts,
            'patents': patents,
            'insurances': insurances,
            'exportInfo': {
                'type': export_type,
                'exportedAt': beijing_now(),
                'exporter': g.username,
                'totalContracts': len(contracts),
                'totalPatents': len(patents),
                'totalInsurances': len(insurances),
            }
        }
        write_log(action='导出数据', module='数据导出',
                  detail=f'JSON导出: 合同{len(contracts)}条, 专利{len(patents)}条, 车险{len(insurances)}条',
                  user_id=g.user_id, username=g.username)
        return jsonify({'success': True, 'data': result})

    # Excel (.xlsx) 导出
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({'success': False, 'message': '服务器未安装 openpyxl，请联系管理员'}), 500

    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def _write_sheet(ws, rows, title):
        if not rows:
            ws.append([f'{title}：无数据'])
            return
        # 写表头
        headers = list(rows[0].keys())
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        # 写数据
        for row_idx, row in enumerate(rows, 2):
            for col_idx, h in enumerate(headers, 1):
                val = row.get(h, '')
                if val is None:
                    val = ''
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
        # 自动列宽
        for col_idx, h in enumerate(headers, 1):
            max_len = len(str(h))
            for row_idx in range(2, len(rows) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

    ws1 = wb.create_sheet('合同数据')
    _write_sheet(ws1, contracts, '合同数据')
    ws2 = wb.create_sheet('专利数据')
    _write_sheet(ws2, patents, '专利数据')
    ws3 = wb.create_sheet('车险数据')
    _write_sheet(ws3, insurances, '车险数据')

    # 保存文件到 downloads 目录
    export_dir = _download_dir()
    os.makedirs(export_dir, exist_ok=True)
    timestamp = beijing_now().replace(':', '-').replace(' ', '_')
    filename = f'export_{timestamp}.xlsx'
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    host_path = _host_path(filepath)

    write_log(action='导出数据', module='数据导出',
              detail=f'Excel导出: {filename} (合同{len(contracts)}条, 专利{len(patents)}条, 车险{len(insurances)}条)',
              user_id=g.user_id, username=g.username)

    return jsonify({
        'success': True,
        'data': {
            'filename': filename,
            'filepath': host_path,
            'exportInfo': {
                'type': export_type,
                'exportedAt': beijing_now(),
                'exporter': g.username,
                'totalContracts': len(contracts),
                'totalPatents': len(patents),
                'totalInsurances': len(insurances),
            }
        },
        'message': f'导出成功！文件已保存至: {host_path}'
    })
